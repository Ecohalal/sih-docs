"""Parser unificado FM 7.8.x → CSV de certificados FAMBRAS.

Lê 3 planilhas:
  - FM 7.8.1 (Industrial ativos)
  - FM 7.8.2 (3 abas: Frigorífico ativos, Cancelados+Suspensos, Vencidos)
  - SUSPENSOS_E_CANCELADOS (3 abas relevantes)

Produz: fm_certificados_consolidado_20260528.csv com colunas canônicas.
"""
import os, re, csv, unicodedata
from datetime import datetime, timedelta
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb

DIR = r"C:\Projetos\Ecohalal\fambras-references-2026-04\_emails-fonte\Qualidade\Lista de clientes"
OUT_DIR = r"C:\Users\ecotrace\OneDrive - ECOTRACE TECNOLOGIA DA INFORMACAO SA\Área de Trabalho\Integra"
OUT_FILE = os.path.join(OUT_DIR, "fm_certificados_consolidado_20260528.csv")
TS = '20260528'

# Header → canonical mapping (chaves são canonical, valores são sinônimos parciais)
HEADER_SYNS = {
    'razao_social': ['nomedaempresarazaosocial', 'nomedaempresacompanyname'],
    'cnpj': ['numeroderegistrocnpjregistrationnumber', 'cnpjregistrationnumber', 'numeroderegistrocnpj'],
    'sif': ['registrosanitariosifsanitaryregistration', 'registrosanitariosif'],
    'endereco': ['enderecoaddress', 'localizacaolocation'],
    'pais': ['paiscountry'],
    'cat_gso': ['categoriagso20551', 'categoriagsouae20551'],
    'cat_smiic': ['categoriaoicsmiic'],
    'categoria_unica': ['categoriacategory'],
    'produto_pt': ['tipodeproduto'],
    'produto_en': ['producttype', 'tipodeprodutoproducttype'],
    'tipos_animais': ['tiposdeanimaisavesoubovinos'],
    'escopo': ['completeescope', 'escopocompletocompletescope', 'escoposcope'],
    'cert_num': ['numerodocertificadocertificatenumber', 'numerodocertificado'],
    'cert_emissao': ['emissao', 'emissaodocertificado', 'emissaodocertificadocertificateissue'],
    'cert_validade': ['validadedocertificado', 'validadedocertificadocertificatevalidity'],
    'reconhece_golfo': ['reconhecimentoincluigolfo', 'reconhecimentoincluigolfogulfrecognition'],
    'normas': ['normasacreditadas', 'normasacreditadasaccreditedstandards'],
    'plataformas': ['cadastronasplataformashaksis', 'cadastronasplataformashaksisbpjph'],
    'status': ['statusdocertificadocertificatestatus', 'statusdocertificate'],
    'motivo': ['motivodocancelamento', 'motivodocancelamentothereason', 'motivodocancelamentoousubstituicao'],
    'data_exclusao': ['datadaexclusaoexclusiondate'],
}

CANONICAL_FIELDS = [
    'origem_arquivo', 'origem_aba', 'origem_linha',
    'razao_social', 'cnpj', 'cnpj_root', 'sif', 'endereco', 'pais',
    'cat_gso', 'cat_smiic', 'categoria_unica',
    'produto_pt', 'produto_en', 'tipos_animais', 'escopo',
    'cert_num', 'cert_empresa_code', 'cert_unidade_code', 'cert_anomes', 'cert_seq', 'cert_revisao', 'cert_pais_code',
    'cert_emissao', 'cert_validade',
    'reconhece_golfo', 'normas', 'plataformas',
    'status', 'status_canonical', 'motivo', 'data_exclusao',
]


def norm_text(s):
    if s is None: return ''
    s = str(s).lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'[\s\W_]+', '', s)


def map_header(h):
    nh = norm_text(h)
    if not nh: return None
    for canonical, syns in HEADER_SYNS.items():
        for syn in syns:
            if syn and syn in nh:
                return canonical
    return None


def coerce_date(v):
    if v is None or v == '':
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=int(v))
            return d.strftime('%Y-%m-%d')
        except Exception:
            return str(v)
    s = str(v).strip()
    # ISO
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m: return m.group(1)
    # Brazilian DD/MM/YYYY
    m = re.search(r'(\d{1,2})/\s*(\d{1,2})/\s*(\d{4})', s)
    if m:
        d, mo, y = m.groups()
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return s
    # DD.MM.YYYY
    m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2: y = '20' + y
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return s
    return s


def normalize_cnpj(v):
    if v is None: return ''
    s = str(v).strip()
    # remover sufixos do tipo "\nManufactured..."
    s = s.split('\n')[0].strip()
    s = re.sub(r'\D', '', s)
    return s


CERT_RE = re.compile(r'^([A-Z]{2,5})\.([A-Z]{2,5})\.(\d{4})\.(\d{2,5})(?:\.(\d+))?\.([A-Z]{2,4})\.?$', re.IGNORECASE)


def parse_cert_num(s):
    out = {}
    if not s: return out
    s_clean = str(s).strip().rstrip('.').replace(' ', '')
    out['cert_num'] = s_clean
    m = CERT_RE.match(s_clean)
    if m:
        emp, unit, anomes, seq, rev, pais = m.groups()
        out.update({
            'cert_empresa_code': emp.upper(),
            'cert_unidade_code': unit.upper(),
            'cert_anomes': anomes,
            'cert_seq': seq,
            'cert_revisao': rev or '',
            'cert_pais_code': pais.upper(),
        })
    return out


def status_canonical(s):
    if not s: return 'unknown'
    t = norm_text(s)
    if 'ativoactive' in t or t.startswith('ativo') or t.startswith('active'):
        return 'ativo'
    if 'cancelado' in t and ('substitu' in t or 'renov' in t):
        return 'cancelado_substituido'
    if 'cancelado' in t:
        return 'cancelado'
    if 'suspenso' in t:
        return 'suspenso'
    if 'vencido' in t or 'expirado' in t:
        return 'vencido'
    return 'outro'


def find_header_row(rows, min_non_null=3):
    """Encontra a primeira linha com pelo menos min_non_null células com valor."""
    for idx, r in enumerate(rows):
        non_null = sum(1 for v in r if v not in (None, '', 0))
        if non_null >= min_non_null:
            # Verifica que parece header (contém pelo menos um nome mapeado)
            for v in r:
                if map_header(v):
                    return idx
    return None


def cnpj_root(c):
    if len(c) >= 8 and c.isdigit():
        return c[:8]
    return ''


def process_row(values, col_map, source):
    record = {f: '' for f in CANONICAL_FIELDS}
    record.update(source)
    for col_idx, canonical in col_map.items():
        if col_idx >= len(values): continue
        v = values[col_idx]
        if v is None: v = ''
        if canonical in ('cert_emissao', 'cert_validade', 'data_exclusao'):
            record[canonical] = coerce_date(v)
        elif canonical == 'cnpj':
            record['cnpj'] = normalize_cnpj(v)
        elif canonical == 'cert_num':
            parsed = parse_cert_num(v)
            for k, val in parsed.items():
                record[k] = val
        else:
            record[canonical] = str(v).strip() if v else ''
    # derived
    record['cnpj_root'] = cnpj_root(record['cnpj'])
    record['status_canonical'] = status_canonical(record.get('status'))
    return record


def process_xlsx(path, target_sheets=None):
    print(f"\n>>> XLSX: {os.path.basename(path)}")
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sh in wb.sheetnames:
        if target_sheets and sh not in target_sheets:
            continue
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = find_header_row(rows)
        if header_idx is None:
            print(f"  ABA '{sh}': sem header detectado, pulando")
            continue
        header = rows[header_idx]
        col_map = {}
        for ci, cell in enumerate(header):
            mapped = map_header(cell)
            if mapped:
                col_map[ci] = mapped
        if not col_map:
            print(f"  ABA '{sh}': nenhuma coluna canônica encontrada")
            continue
        mapped_set = set(col_map.values())
        print(f"  ABA '{sh}': header L{header_idx+1}, {len(col_map)} colunas mapeadas: {sorted(mapped_set)}")
        cnt = 0
        for ri, row in enumerate(rows[header_idx+1:], start=header_idx+2):
            # pular linhas vazias
            if all(v in (None, '') for v in row):
                continue
            source = {
                'origem_arquivo': os.path.basename(path),
                'origem_aba': sh,
                'origem_linha': str(ri),
            }
            rec = process_row(list(row), col_map, source)
            if not rec.get('razao_social') and not rec.get('cert_num'):
                continue
            out.append(rec)
            cnt += 1
        print(f"    {cnt} registros extraídos")
    wb.close()
    return out


def process_xlsb(path, target_sheets=None):
    print(f"\n>>> XLSB: {os.path.basename(path)}")
    out = []
    with open_xlsb(path) as wb:
        for sh in wb.sheets:
            if target_sheets and sh not in target_sheets:
                continue
            with wb.get_sheet(sh) as sheet:
                rows = []
                for row in sheet.rows():
                    rows.append([c.v for c in row])
                header_idx = find_header_row(rows)
                if header_idx is None:
                    print(f"  ABA '{sh}': sem header detectado, pulando")
                    continue
                header = rows[header_idx]
                col_map = {}
                for ci, cell in enumerate(header):
                    mapped = map_header(cell)
                    if mapped:
                        col_map[ci] = mapped
                if not col_map:
                    print(f"  ABA '{sh}': nenhuma coluna canônica encontrada")
                    continue
                mapped_set = set(col_map.values())
                print(f"  ABA '{sh}': header L{header_idx+1}, {len(col_map)} colunas mapeadas: {sorted(mapped_set)}")
                cnt = 0
                for ri, row in enumerate(rows[header_idx+1:], start=header_idx+2):
                    if all(v in (None, '') for v in row):
                        continue
                    source = {
                        'origem_arquivo': os.path.basename(path),
                        'origem_aba': sh,
                        'origem_linha': str(ri),
                    }
                    rec = process_row(row, col_map, source)
                    if not rec.get('razao_social') and not rec.get('cert_num'):
                        continue
                    out.append(rec)
                    cnt += 1
                print(f"    {cnt} registros extraídos")
    return out


# --- Run ---
records = []
records += process_xlsx(os.path.join(DIR, "FM 7.8.1 - CERTIFICATED PRODUCTS INDUSTRIAL_ATIVOS - 14.05.2026.xlsx"))
records += process_xlsb(os.path.join(DIR, "FM 7.8.2 - CERTIFICATED PRODUCTS LIST SLAUGHTERHOUSE  - 08.04.2026.xlsb"))
records += process_xlsx(
    os.path.join(DIR, "SUSPENSOS E CANCELADOS_IND E FRIG_26.03.2026.xlsx"),
    target_sheets={'SITE', 'SITE_HISTÓRICO COMPLETO', 'EXCLUIDOS DA LISTA - SITE'},
)

print(f"\nTotal registros: {len(records)}")

# --- Write CSV ---
os.makedirs(OUT_DIR, exist_ok=True)
with open(OUT_FILE, 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=CANONICAL_FIELDS, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    w.writeheader()
    for r in records:
        w.writerow(r)
print(f"Salvo: {OUT_FILE} ({os.path.getsize(OUT_FILE)} bytes)")

# --- Summary ---
from collections import Counter
sep = '=' * 70
print(f"\n{sep}\nSTATUS CANONICAL\n{sep}")
for k, v in Counter(r['status_canonical'] for r in records).most_common():
    print(f"  {k}: {v}")

print(f"\n{sep}\nORIGEM (arquivo/aba)\n{sep}")
for k, v in Counter((r['origem_arquivo'], r['origem_aba']) for r in records).most_common():
    print(f"  [{v:4}]  {k[0]}  >  {k[1]}")

print(f"\n{sep}\nCERT_NUM PARSING\n{sep}")
parsed_ok = sum(1 for r in records if r.get('cert_empresa_code'))
print(f"  parsed_ok: {parsed_ok}")
print(f"  parse_falhou (cert_num present mas regex nao bateu): {sum(1 for r in records if r.get('cert_num') and not r.get('cert_empresa_code'))}")
print(f"  cert_num vazio: {sum(1 for r in records if not r.get('cert_num'))}")

print(f"\n{sep}\nPAÍS NO CERT NUMBER\n{sep}")
for k, v in Counter(r['cert_pais_code'] for r in records if r.get('cert_pais_code')).most_common(10):
    print(f"  {k}: {v}")

print(f"\n{sep}\nCNPJ DUPLICATION (mesma planta tem múltiplos certs)\n{sep}")
cnpjs = [r['cnpj'] for r in records if r.get('cnpj')]
cnpj_counter = Counter(cnpjs)
print(f"  CNPJs únicos: {len(cnpj_counter)}")
print(f"  CNPJs com >1 cert: {sum(1 for c, n in cnpj_counter.items() if n > 1)}")
print(f"  Top 5 CNPJs por nº de certs:")
for c, n in cnpj_counter.most_common(5):
    print(f"    {c}: {n} certs")

print(f"\n{sep}\nCROSS-CHECK: cert_anomes (10 mais recentes vs 10 mais antigos)\n{sep}")
anomes = sorted({r['cert_anomes'] for r in records if r.get('cert_anomes')})
if anomes:
    print(f"  Range: {anomes[0]} → {anomes[-1]}  (total {len(anomes)} anomes únicos)")
