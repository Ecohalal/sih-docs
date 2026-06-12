"""Probe das planilhas FM 7.8.x da FAMBRAS — só lê estrutura e amostra.
Não modifica nada. Identifica abas, headers, e 5 linhas de cada.
"""
import os, sys
from openpyxl import load_workbook
try:
    from pyxlsb import open_workbook as open_xlsb
    HAS_XLSB = True
except Exception:
    HAS_XLSB = False

DIR = r"C:\Projetos\Ecohalal\fambras-references-2026-04\_emails-fonte\Qualidade\Lista de clientes"

FILES = [
    "FM 7.8.1 - CERTIFICATED PRODUCTS INDUSTRIAL_ATIVOS - 14.05.2026.xlsx",
    "FM 7.8.2 - CERTIFICATED PRODUCTS LIST SLAUGHTERHOUSE  - 08.04.2026.xlsb",
    "SUSPENSOS E CANCELADOS_IND E FRIG_26.03.2026.xlsx",
    "X FM 7.8.1.S - INDUSTRIAL - SITE_16.04.2026.xlsx",
    "X FM 7.8.2.S - FRIGORIFICO - SITE_08.04.2026.xlsx",
]

SEP = "=" * 80


def show_xlsx(path):
    print(SEP); print(f"FILE: {os.path.basename(path)}"); print(SEP)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERRO ao abrir: {e}")
        return
    for sh_name in wb.sheetnames:
        ws = wb[sh_name]
        # Conta linhas/cols
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"\n  ABA: '{sh_name}'  | linhas={max_row}  cols={max_col}")
        rows = list(ws.iter_rows(max_row=8, values_only=True))
        if not rows:
            print("    (vazia)")
            continue
        # Encontrar a primeira linha 'header-like' (>=3 não-nulos)
        header_idx = None
        for idx, r in enumerate(rows):
            non_null = sum(1 for v in r if v not in (None, ''))
            if non_null >= 3:
                header_idx = idx
                break
        if header_idx is None:
            print("    (sem header óbvio nas primeiras 8 linhas)")
            for i, r in enumerate(rows):
                print(f"    L{i+1}: {r[:15]}")
            continue
        header = rows[header_idx]
        # Truncar pra não estourar
        clean_header = [str(c)[:30] if c is not None else '' for c in header[:25]]
        print(f"    HEADER (L{header_idx+1}): {clean_header}")
        print(f"    SAMPLE (próximas 3 linhas):")
        for i, r in enumerate(rows[header_idx+1:header_idx+4], start=header_idx+2):
            clean = [str(c)[:30] if c is not None else '' for c in r[:25]]
            print(f"      L{i}: {clean}")
    wb.close()


def show_xlsb(path):
    print(SEP); print(f"FILE (xlsb): {os.path.basename(path)}"); print(SEP)
    if not HAS_XLSB:
        print("  pyxlsb não instalado")
        return
    try:
        with open_xlsb(path) as wb:
            sheet_names = wb.sheets
            for sh_name in sheet_names:
                with wb.get_sheet(sh_name) as sheet:
                    print(f"\n  ABA: '{sh_name}'")
                    rows = []
                    for i, row in enumerate(sheet.rows()):
                        rows.append([c.v for c in row])
                        if i >= 7:
                            break
                    if not rows:
                        print("    (vazia)")
                        continue
                    # Header idx
                    header_idx = None
                    for idx, r in enumerate(rows):
                        non_null = sum(1 for v in r if v not in (None, ''))
                        if non_null >= 3:
                            header_idx = idx
                            break
                    if header_idx is None:
                        for i, r in enumerate(rows):
                            print(f"    L{i+1}: {r[:15]}")
                        continue
                    header = rows[header_idx]
                    clean_header = [str(c)[:30] if c is not None else '' for c in header[:25]]
                    print(f"    HEADER (L{header_idx+1}): {clean_header}")
                    print(f"    SAMPLE (próximas 3 linhas):")
                    for i, r in enumerate(rows[header_idx+1:header_idx+4], start=header_idx+2):
                        clean = [str(c)[:30] if c is not None else '' for c in r[:25]]
                        print(f"      L{i}: {clean}")
    except Exception as e:
        print(f"  ERRO: {e}")


for fn in FILES:
    path = os.path.join(DIR, fn)
    if not os.path.exists(path):
        print(f"NOT FOUND: {path}")
        continue
    if path.lower().endswith(".xlsb"):
        show_xlsb(path)
    else:
        show_xlsx(path)
